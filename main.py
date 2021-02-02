import openpyxl
from queue import Queue
def fetchdataindict():
    path = './hierarchy_case.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    employee = {}
    manager = {}
    convertedjson = {}
    headlines = []
    ceoid = ''
    
    for i in range(0, sheet_obj.max_row-1):
        employee_details = {}
        for j in range(1, sheet_obj.max_column):
            if i==0:
                headlines.append(sheet_obj.cell(row=i+1, column=j+1).value)
                continue
            else:
                employee_details[headlines[j-1]] = sheet_obj.cell(row=i+1, column=j+1).value
        if (not employee_details):
            continue
        if (employee_details[headlines[-1]]==None):
            ceoid = employee_details[headlines[0]]
        if employee_details[headlines[-1]] in manager.keys():
            reporties = manager[employee_details[headlines[-1]]]
            if(reporties!=None):
                reporties.append(employee_details[headlines[0]])
                manager[employee_details[headlines[-1]]] = reporties
        else:
            reporties = [employee_details[headlines[0]]]
            manager[employee_details[headlines[-1]]] = reporties
        employee[employee_details[headlines[0]]] = employee_details

    onlymanagers = getonlymanagers(employee, manager, ceoid)
    convertedjson['reportees'] = manager[onlymanagers[-1]]
    convertedjson = convertjson(employee, manager, ceoid, onlymanagers, len(onlymanagers)-1, convertedjson)
    return convertedjson['reportees']


def getonlymanagers(employees, manager, ceoid):
    queue = Queue()
    queue.put(None)
    onlymanagers = []
    while(queue.empty()==False):
        queue_element = queue.get()
        if queue_element not in manager.keys():
            continue
        for i in manager[queue_element]:
            queue.put(i)
        onlymanagers.append(queue_element)

    return onlymanagers


def convertjson(employees, manager, ceoid, onlymanagers, N, convertedjson):
    if(N<1): 
        return convertedjson
    managerdata = {}
    managerdata['employeeID'] = employees[onlymanagers[N]]['EMPLOYEE_ID']
    managerdata['name'] = employees[onlymanagers[N]]['NAME']
    managerdata['reportees'] = [convertedjson['reportees']]
    if(N==len(onlymanagers)-1):
        convertidtonames = []
        for i in convertedjson['reportees']:
            leftoutemployees = {
                'employeeID': employees[i]['EMPLOYEE_ID'], 'name': employees[i]['NAME']}
            convertidtonames.append(leftoutemployees)
        managerdata['reportees'] = convertidtonames
    else:
        for i in manager[managerdata['employeeID']]:
            if ('employeeID' in convertedjson.keys()) and (i == manager[convertedjson['reportees']['employeeID']]):
                continue
            leftoutemployees = {
                'employeeID': employees[i]['EMPLOYEE_ID'], 'name': employees[i]['NAME']}
            managerdata['reportees'].insert(0,leftoutemployees)
    convertedjson = {'reportees': managerdata}
    return convertjson(employees, manager, ceoid, onlymanagers, N-1, convertedjson)

output = fetchdataindict()
print(output)
